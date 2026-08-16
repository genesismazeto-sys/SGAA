import base64
import io
import json
import logging
import os
import sqlite3
import sys
import uuid
from types import SimpleNamespace

import openpyxl
import pytest


BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if BASE not in sys.path:
    sys.path.insert(0, BASE)

import main
from app import db as app_db_module
from app.versioning import resolver as versioning_resolver
from app.versioning import shadow_reads as versioning_shadow_reads
from tests.versioned_test_support import isolated_versioned_app_env


@pytest.fixture()
def shadow_read_env(tmp_path):
    with isolated_versioned_app_env(tmp_path, "shadow_read.db") as env:
        yield {"client": env["client"], "db_path": env["db_path"]}


def _set_admin_session(client):
    with client.session_transaction() as sess:
        sess["user_id"] = 1
        sess["user_type"] = "admin"
        sess["user_name"] = "Administrador"


def _set_aluno_session(client, *, user_id, user_name):
    with client.session_transaction() as sess:
        sess["user_id"] = user_id
        sess["user_type"] = "aluno"
        sess["user_name"] = user_name


def _seed_student_in_turma(codigo_turma="PPA-T11"):
    token = uuid.uuid4().hex[:8]
    email = f"shadow.read.{token}@example.com"
    nome = f"Aluno Shadow {token}"

    with main.app.app_context():
        conn = main.get_db_connection()
        turma = conn.execute(
            "SELECT id, codigo FROM turmas WHERE codigo = ?",
            (codigo_turma,),
        ).fetchone()
        assert turma is not None

        conn.execute(
            "INSERT INTO usuarios (nome, email, senha, tipo, nivel_acesso) VALUES (?, ?, ?, ?, ?)",
            (nome, email, main.hash_password("aluno123"), "aluno", "usuario"),
        )
        usuario_id = conn.execute(
            "SELECT id FROM usuarios WHERE email = ?",
            (email,),
        ).fetchone()["id"]

        conn.execute(
            "INSERT INTO alunos (usuario_id, nome, matricula, email, turma_id, status) VALUES (?, ?, ?, ?, ?, ?)",
            (usuario_id, nome, f"{codigo_turma}.{token}", email, turma["id"], "Ativo"),
        )
        aluno_id = conn.execute(
            "SELECT id FROM alunos WHERE usuario_id = ?",
            (usuario_id,),
        ).fetchone()["id"]
        conn.commit()

    return {
        "usuario_id": usuario_id,
        "aluno_id": aluno_id,
        "nome": nome,
        "turma_codigo": codigo_turma,
    }


def _fetch_req(aluno_id, nome_evento):
    with main.app.app_context():
        conn = main.get_db_connection()
        return conn.execute(
            """
            SELECT id, atividade_id, atividade_versao_id, regra_snapshot_json, codigo_normativo_snapshot, status
              FROM requisicoes
             WHERE aluno_id = ? AND nome_evento = ?
             ORDER BY id DESC
             LIMIT 1
            """,
            (aluno_id, nome_evento),
        ).fetchone()


def _fetch_req_by_id(req_id):
    with main.app.app_context():
        conn = main.get_db_connection()
        return conn.execute(
            """
            SELECT id, aluno_id, atividade_id, atividade_versao_id, regra_snapshot_json,
                   codigo_normativo_snapshot, status, observacao
              FROM requisicoes
             WHERE id = ?
            """,
            (req_id,),
        ).fetchone()


def _resolve_for_student(aluno_id, atividade_id_legacy):
    with main.app.app_context():
        conn = main.get_db_connection()
        return main.resolver_versao_por_aluno(
            conn,
            aluno_id=aluno_id,
            atividade_id_legacy=atividade_id_legacy,
            strict_legacy_scope=True,
        )


def _load_snapshot(req_row):
    assert req_row["regra_snapshot_json"] is not None
    return json.loads(req_row["regra_snapshot_json"])


def _insert_pending_req(aluno_id, atividade_id, *, nome_evento=None):
    nome_evento = nome_evento or f"Req Manual {uuid.uuid4().hex[:6]}"
    with main.app.app_context():
        conn = main.get_db_connection()
        cur = conn.execute(
            """
            INSERT INTO requisicoes (
                aluno_id, atividade_id, data_solicitacao, data_evento,
                horas_solicitadas, nome_evento, status, observacao
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                aluno_id,
                atividade_id,
                "2026-05-30 10:00:00",
                "2026-05-24",
                4,
                nome_evento,
                "Pendente",
                "requisicao manual de teste",
            ),
        )
        req_id = cur.lastrowid
        conn.commit()
    return req_id


def _log_recorder(monkeypatch):
    info_logs = []
    warning_logs = []
    exception_logs = []

    def _capture_info(msg, *args):
        info_logs.append(msg % args if args else msg)

    def _capture_warning(msg, *args):
        warning_logs.append(msg % args if args else msg)

    def _capture_exception(msg, *args):
        exception_logs.append(msg % args if args else msg)

    monkeypatch.setattr(main.logger, "info", _capture_info)
    monkeypatch.setattr(main.logger, "warning", _capture_warning)
    monkeypatch.setattr(main.logger, "exception", _capture_exception)
    return info_logs, warning_logs, exception_logs


def test_shadow_read_student_create_calls_resolver_when_flag_on(shadow_read_env, monkeypatch):
    env = shadow_read_env
    client = env["client"]
    db_path = env["db_path"]
    seed = _seed_student_in_turma()
    _set_aluno_session(client, user_id=seed["usuario_id"], user_name=seed["nome"])
    info_logs, _, _ = _log_recorder(monkeypatch)
    monkeypatch.setenv("SGAA_VERSIONED_RESOLVER_SHADOW_READ", "1")

    event_name = f"Evento Shadow Aluno {uuid.uuid4().hex[:6]}"
    calls = []

    def fake_resolver(conn, *, aluno_id, atividade_id_legacy, strict_legacy_scope=True):
        calls.append((aluno_id, atividade_id_legacy, strict_legacy_scope))
        aux = sqlite3.connect(db_path)
        aux.row_factory = sqlite3.Row
        try:
            row = aux.execute(
                """
                SELECT atividade_id, atividade_versao_id, regra_snapshot_json, codigo_normativo_snapshot
                  FROM requisicoes
                 WHERE aluno_id = ? AND nome_evento = ?
                 ORDER BY id DESC
                 LIMIT 1
                """,
                (seed["aluno_id"], event_name),
            ).fetchone()
        finally:
            aux.close()
        if row is not None:
            assert row["atividade_id"] == 1
            assert row["atividade_versao_id"] is not None
            assert row["regra_snapshot_json"] is not None
            assert row["codigo_normativo_snapshot"] is not None
        assert strict_legacy_scope is True
        return {
            "status": "resolved",
            "atividade_versao_id": 29,
            "atividade_base_id": 1,
            "codigo_normativo": "AAC-rev6",
            "eixo": "AAC",
            "matriz_id_efetiva": 2,
            "legacy_scope_ok": True,
            "warnings": [],
            "reason": "ok",
        }

    monkeypatch.setattr(versioning_resolver, "resolver_versao_por_aluno", fake_resolver)

    response = client.post(
        "/aluno/nova-requisicao",
        data={
            "atividade_id": "1",
            "nome_evento": event_name,
            "data_evento": "2026-05-24",
            "horas_solicitadas": "4",
            "observacao": "shadow read aluno",
        },
        follow_redirects=False,
    )

    assert response.status_code in (302, 303)
    assert calls == [(seed["aluno_id"], 1, True), (seed["aluno_id"], 1, True)]
    req = _fetch_req(seed["aluno_id"], event_name)
    assert req is not None
    assert req["atividade_id"] == 1
    assert req["atividade_versao_id"] == 29
    assert req["regra_snapshot_json"] is not None
    assert req["codigo_normativo_snapshot"] == "AAC-rev6"
    assert any("event=versioned_resolver_shadow_read" in log and "origin=aluno_create" in log for log in info_logs)


def test_mandatory_student_snapshot_rejects_resolver_error(shadow_read_env, monkeypatch):
    env = shadow_read_env
    client = env["client"]
    seed = _seed_student_in_turma()
    _set_aluno_session(client, user_id=seed["usuario_id"], user_name=seed["nome"])
    _, _, exception_logs = _log_recorder(monkeypatch)
    monkeypatch.setenv("SGAA_VERSIONED_RESOLVER_SHADOW_READ", "1")

    event_name = f"Evento Shadow Aluno Erro {uuid.uuid4().hex[:6]}"

    def raising_resolver(conn, *, aluno_id, atividade_id_legacy, strict_legacy_scope=True):
        raise RuntimeError("boom aluno")

    monkeypatch.setattr(versioning_resolver, "resolver_versao_por_aluno", raising_resolver)

    response = client.post(
        "/aluno/nova-requisicao",
        data={
            "atividade_id": "1",
            "nome_evento": event_name,
            "data_evento": "2026-05-24",
            "horas_solicitadas": "6",
            "observacao": "shadow read aluno erro",
        },
        follow_redirects=False,
    )

    assert response.status_code == 200
    req = _fetch_req(seed["aluno_id"], event_name)
    assert req is None
    assert not any("event=versioned_resolver_shadow_read" in log for log in exception_logs)


def test_student_create_does_not_skip_mandatory_resolver_when_shadow_flag_off(shadow_read_env, monkeypatch):
    env = shadow_read_env
    client = env["client"]
    seed = _seed_student_in_turma()
    _set_aluno_session(client, user_id=seed["usuario_id"], user_name=seed["nome"])
    monkeypatch.delenv("SGAA_VERSIONED_RESOLVER_SHADOW_READ", raising=False)

    def forbidden_resolver(conn, *, aluno_id, atividade_id_legacy, strict_legacy_scope=True):
        raise AssertionError("mandatory resolver probe")

    monkeypatch.setattr(versioning_resolver, "resolver_versao_por_aluno", forbidden_resolver)

    event_name = f"Evento Shadow Aluno Off {uuid.uuid4().hex[:6]}"
    response = client.post(
        "/aluno/nova-requisicao",
        data={
            "atividade_id": "1",
            "nome_evento": event_name,
            "data_evento": "2026-05-24",
            "horas_solicitadas": "2",
            "observacao": "flag off aluno",
        },
        follow_redirects=False,
    )

    assert response.status_code == 200
    req = _fetch_req(seed["aluno_id"], event_name)
    assert req is None


def test_shadow_read_admin_create_calls_resolver_when_flag_on(shadow_read_env, monkeypatch):
    env = shadow_read_env
    client = env["client"]
    db_path = env["db_path"]
    seed = _seed_student_in_turma()
    _set_admin_session(client)
    info_logs, _, _ = _log_recorder(monkeypatch)
    monkeypatch.setenv("SGAA_VERSIONED_RESOLVER_SHADOW_READ", "1")

    event_name = f"Evento Shadow Admin {uuid.uuid4().hex[:6]}"
    calls = []

    def fake_resolver(conn, *, aluno_id, atividade_id_legacy, strict_legacy_scope=True):
        calls.append((aluno_id, atividade_id_legacy, strict_legacy_scope))
        aux = sqlite3.connect(db_path)
        aux.row_factory = sqlite3.Row
        try:
            row = aux.execute(
                """
                SELECT atividade_id, atividade_versao_id, regra_snapshot_json, codigo_normativo_snapshot
                  FROM requisicoes
                 WHERE aluno_id = ? AND nome_evento = ?
                 ORDER BY id DESC
                 LIMIT 1
                """,
                (seed["aluno_id"], event_name),
            ).fetchone()
        finally:
            aux.close()
        if row is not None:
            assert row["atividade_id"] == 1
            assert row["atividade_versao_id"] is not None
            assert row["regra_snapshot_json"] is not None
            assert row["codigo_normativo_snapshot"] is not None
        assert strict_legacy_scope is True
        return {
            "status": "resolved",
            "atividade_versao_id": 29,
            "atividade_base_id": 1,
            "codigo_normativo": "AAC-rev6",
            "eixo": "AAC",
            "matriz_id_efetiva": 2,
            "legacy_scope_ok": True,
            "warnings": [],
            "reason": "ok",
        }

    monkeypatch.setattr(versioning_resolver, "resolver_versao_por_aluno", fake_resolver)

    response = client.post(
        "/admin/requisicoes/nova",
        data={
            "aluno_id": str(seed["aluno_id"]),
            "atividade_id": "1",
            "nome_evento": event_name,
            "data_evento": "2026-05-24",
            "horas_solicitadas": "5",
            "observacao": "shadow read admin",
        },
        follow_redirects=False,
    )

    assert response.status_code in (302, 303)
    assert calls == [(seed["aluno_id"], 1, True), (seed["aluno_id"], 1, True)]
    req = _fetch_req(seed["aluno_id"], event_name)
    assert req is not None
    assert req["atividade_id"] == 1
    assert req["atividade_versao_id"] == 29
    assert req["regra_snapshot_json"] is not None
    assert req["codigo_normativo_snapshot"] == "AAC-rev6"
    assert any("event=versioned_resolver_shadow_read" in log and "origin=admin_create" in log for log in info_logs)


def test_mandatory_admin_snapshot_rejects_resolver_error(shadow_read_env, monkeypatch):
    env = shadow_read_env
    client = env["client"]
    seed = _seed_student_in_turma()
    _set_admin_session(client)
    _, _, exception_logs = _log_recorder(monkeypatch)
    monkeypatch.setenv("SGAA_VERSIONED_RESOLVER_SHADOW_READ", "1")

    event_name = f"Evento Shadow Admin Erro {uuid.uuid4().hex[:6]}"

    def raising_resolver(conn, *, aluno_id, atividade_id_legacy, strict_legacy_scope=True):
        raise RuntimeError("boom admin")

    monkeypatch.setattr(versioning_resolver, "resolver_versao_por_aluno", raising_resolver)

    response = client.post(
        "/admin/requisicoes/nova",
        data={
            "aluno_id": str(seed["aluno_id"]),
            "atividade_id": "1",
            "nome_evento": event_name,
            "data_evento": "2026-05-24",
            "horas_solicitadas": "7",
            "observacao": "shadow read admin erro",
        },
        follow_redirects=False,
    )

    assert response.status_code in (302, 303)
    req = _fetch_req(seed["aluno_id"], event_name)
    assert req is None
    assert not any("event=versioned_resolver_shadow_read" in log for log in exception_logs)


def test_admin_create_does_not_skip_mandatory_resolver_when_shadow_flag_off(shadow_read_env, monkeypatch):
    env = shadow_read_env
    client = env["client"]
    seed = _seed_student_in_turma()
    _set_admin_session(client)
    monkeypatch.delenv("SGAA_VERSIONED_RESOLVER_SHADOW_READ", raising=False)

    def forbidden_resolver(conn, *, aluno_id, atividade_id_legacy, strict_legacy_scope=True):
        raise AssertionError("mandatory resolver probe")

    monkeypatch.setattr(versioning_resolver, "resolver_versao_por_aluno", forbidden_resolver)

    event_name = f"Evento Shadow Admin Off {uuid.uuid4().hex[:6]}"
    response = client.post(
        "/admin/requisicoes/nova",
        data={
            "aluno_id": str(seed["aluno_id"]),
            "atividade_id": "1",
            "nome_evento": event_name,
            "data_evento": "2026-05-24",
            "horas_solicitadas": "3",
            "observacao": "flag off admin",
        },
        follow_redirects=False,
    )

    assert response.status_code in (302, 303)
    req = _fetch_req(seed["aluno_id"], event_name)
    assert req is None


def test_snapshot_write_student_create_writes_aac_snapshot_with_shadow_off(shadow_read_env, monkeypatch):
    env = shadow_read_env
    client = env["client"]
    seed = _seed_student_in_turma()
    _set_aluno_session(client, user_id=seed["usuario_id"], user_name=seed["nome"])
    monkeypatch.setenv("SGAA_VERSIONED_REQUISICAO_SNAPSHOT_WRITE", "1")
    monkeypatch.delenv("SGAA_VERSIONED_RESOLVER_SHADOW_READ", raising=False)

    expected = _resolve_for_student(seed["aluno_id"], 1)
    assert expected["status"] == "resolved"

    event_name = f"Evento Snapshot AAC {uuid.uuid4().hex[:6]}"
    response = client.post(
        "/aluno/nova-requisicao",
        data={
            "atividade_id": "1",
            "nome_evento": event_name,
            "data_evento": "2026-05-24",
            "horas_solicitadas": "4",
            "observacao": "snapshot aluno aac",
        },
        follow_redirects=False,
    )

    assert response.status_code in (302, 303)
    req = _fetch_req(seed["aluno_id"], event_name)
    assert req is not None
    assert req["atividade_id"] == 1
    assert req["atividade_versao_id"] == expected["atividade_versao_id"]
    assert req["codigo_normativo_snapshot"] == expected["codigo_normativo"]
    snapshot = _load_snapshot(req)
    assert snapshot["schema_version"] == "d6.4.0-v1"
    assert snapshot["flow_origin"] == "aluno_create"
    assert snapshot["resolver_status"] == "resolved"
    assert snapshot["atividade_id_legacy"] == 1
    assert snapshot["atividade_versao_id"] == expected["atividade_versao_id"]
    assert snapshot["atividade_base_id"] == expected["atividade_base_id"]
    assert snapshot["codigo_normativo"] == expected["codigo_normativo"]
    assert snapshot["eixo"] == expected["eixo"] == "AAC"
    assert snapshot["legacy_scope_ok"] is True
    assert snapshot["matriz_id_efetiva"] == expected["matriz_id_efetiva"]
    assert snapshot["resolver_warnings"] == []
    assert snapshot["nome_exibivel"]
    assert snapshot["nome_legacy"]
    assert snapshot["tipo_atividade_legacy"] == "Acadêmica Complementar"
    assert snapshot["versao_status"] == "ativa"
    assert "T" in snapshot["snapshot_written_at"]
    assert snapshot["snapshot_written_at"].endswith("Z")
    assert not os.path.exists(versioning_shadow_reads._versioned_shadow_read_dedicated_log_path())


def test_snapshot_write_admin_create_writes_aeu_snapshot_with_shadow_off(shadow_read_env, monkeypatch):
    env = shadow_read_env
    client = env["client"]
    seed = _seed_student_in_turma()
    _set_admin_session(client)
    monkeypatch.setenv("SGAA_VERSIONED_REQUISICAO_SNAPSHOT_WRITE", "1")
    monkeypatch.delenv("SGAA_VERSIONED_RESOLVER_SHADOW_READ", raising=False)

    expected = _resolve_for_student(seed["aluno_id"], 29)
    assert expected["status"] == "resolved"
    assert expected["eixo"] == "AEU"

    event_name = f"Evento Snapshot AEU {uuid.uuid4().hex[:6]}"
    response = client.post(
        "/admin/requisicoes/nova",
        data={
            "aluno_id": str(seed["aluno_id"]),
            "atividade_id": "29",
            "nome_evento": event_name,
            "data_evento": "2026-05-24",
            "horas_solicitadas": "5",
            "observacao": "snapshot admin aeu",
        },
        follow_redirects=False,
    )

    assert response.status_code in (302, 303)
    req = _fetch_req(seed["aluno_id"], event_name)
    assert req is not None
    assert req["atividade_id"] == 29
    assert req["atividade_versao_id"] == expected["atividade_versao_id"]
    assert req["codigo_normativo_snapshot"] == expected["codigo_normativo"]
    snapshot = _load_snapshot(req)
    assert snapshot["flow_origin"] == "admin_create"
    assert snapshot["atividade_id_legacy"] == 29
    assert snapshot["atividade_versao_id"] == expected["atividade_versao_id"]
    assert snapshot["codigo_normativo"] == expected["codigo_normativo"] == "AEU-rev1"
    assert snapshot["eixo"] == "AEU"
    assert snapshot["tipo_atividade_legacy"] == "Extensão Universitária"
    assert snapshot["legacy_scope_ok"] is True
    assert snapshot["resolver_status"] == "resolved"
    assert snapshot["resolver_warnings"] == []
    assert not os.path.exists(versioning_shadow_reads._versioned_shadow_read_dedicated_log_path())


def test_mandatory_student_snapshot_rejects_unresolved_resolver(shadow_read_env, monkeypatch):
    env = shadow_read_env
    client = env["client"]
    seed = _seed_student_in_turma()
    _set_aluno_session(client, user_id=seed["usuario_id"], user_name=seed["nome"])
    info_logs, warning_logs, exception_logs = _log_recorder(monkeypatch)
    monkeypatch.setenv("SGAA_VERSIONED_REQUISICAO_SNAPSHOT_WRITE", "1")
    monkeypatch.delenv("SGAA_VERSIONED_RESOLVER_SHADOW_READ", raising=False)

    def fake_resolver(conn, *, aluno_id, atividade_id_legacy, strict_legacy_scope=True):
        return {
            "status": "legacy_activity_without_base_map",
            "atividade_versao_id": None,
            "atividade_base_id": None,
            "codigo_normativo": None,
            "eixo": None,
            "matriz_id_efetiva": 2,
            "legacy_scope_ok": True,
            "warnings": [],
            "reason": "sem mapeamento para base",
        }

    monkeypatch.setattr(versioning_resolver, "resolver_versao_por_aluno", fake_resolver)

    event_name = f"Evento Snapshot Unresolved {uuid.uuid4().hex[:6]}"
    response = client.post(
        "/aluno/nova-requisicao",
        data={
            "atividade_id": "1",
            "nome_evento": event_name,
            "data_evento": "2026-05-24",
            "horas_solicitadas": "3",
            "observacao": "snapshot unresolved",
        },
        follow_redirects=False,
    )

    assert response.status_code == 200
    req = _fetch_req(seed["aluno_id"], event_name)
    assert req is None
    assert info_logs == []
    assert exception_logs == []
    assert warning_logs == []


def test_snapshot_write_admin_edit_remains_out_of_scope(shadow_read_env, monkeypatch):
    env = shadow_read_env
    client = env["client"]
    seed = _seed_student_in_turma()
    _set_admin_session(client)
    monkeypatch.setenv("SGAA_VERSIONED_REQUISICAO_SNAPSHOT_WRITE", "1")
    monkeypatch.delenv("SGAA_VERSIONED_RESOLVER_SHADOW_READ", raising=False)

    def forbidden_resolver(conn, *, aluno_id, atividade_id_legacy, strict_legacy_scope=True):
        raise AssertionError("resolvedor nao deveria ser chamado em admin_edit")

    monkeypatch.setattr(versioning_resolver, "resolver_versao_por_aluno", forbidden_resolver)

    req_id = _insert_pending_req(seed["aluno_id"], 1, nome_evento="Req edit sem snapshot")
    response = client.post(
        f"/admin/requisicoes/{req_id}/editar",
        data={
            "atividade_id": "1",
            "nome_evento": "Req edit sem snapshot atualizada",
            "horas_solicitadas": "8",
            "data_evento": "2026-05-25",
            "observacao": "edicao sem dual-write",
        },
        follow_redirects=False,
    )

    assert response.status_code in (302, 303)
    req = _fetch_req_by_id(req_id)
    assert req is not None
    assert req["atividade_versao_id"] is None
    assert req["regra_snapshot_json"] is None
    assert req["codigo_normativo_snapshot"] is None


def test_snapshot_write_admin_processar_remains_out_of_scope(shadow_read_env, monkeypatch):
    env = shadow_read_env
    client = env["client"]
    seed = _seed_student_in_turma()
    _set_admin_session(client)
    monkeypatch.setenv("SGAA_VERSIONED_REQUISICAO_SNAPSHOT_WRITE", "1")
    monkeypatch.delenv("SGAA_VERSIONED_RESOLVER_SHADOW_READ", raising=False)

    def forbidden_resolver(conn, *, aluno_id, atividade_id_legacy, strict_legacy_scope=True):
        raise AssertionError("resolvedor nao deveria ser chamado em admin_processar_requisicao")

    monkeypatch.setattr(versioning_resolver, "resolver_versao_por_aluno", forbidden_resolver)

    req_id = _insert_pending_req(seed["aluno_id"], 1, nome_evento="Req process sem snapshot")
    response = client.post(
        f"/admin/processar_requisicao/{req_id}",
        data={
            "status": "Indeferida",
            "observacao": "processamento sem dual-write",
        },
        follow_redirects=False,
    )

    assert response.status_code in (302, 303)
    req = _fetch_req_by_id(req_id)
    assert req is not None
    assert req["status"] == "Indeferida"
    assert req["atividade_versao_id"] is None
    assert req["regra_snapshot_json"] is None
    assert req["codigo_normativo_snapshot"] is None


def test_snapshot_write_admin_importar_remains_out_of_scope(shadow_read_env, monkeypatch):
    env = shadow_read_env
    client = env["client"]
    _set_admin_session(client)
    monkeypatch.setenv("SGAA_VERSIONED_REQUISICAO_SNAPSHOT_WRITE", "1")
    monkeypatch.delenv("SGAA_VERSIONED_RESOLVER_SHADOW_READ", raising=False)

    def forbidden_resolver(conn, *, aluno_id, atividade_id_legacy, strict_legacy_scope=True):
        raise AssertionError("resolvedor nao deveria ser chamado em admin_importar_requisicoes")

    monkeypatch.setattr(versioning_resolver, "resolver_versao_por_aluno", forbidden_resolver)

    with main.app.app_context():
        conn = main.get_db_connection()
        atividade_nome = conn.execute(
            "SELECT nome FROM atividades WHERE id = 1"
        ).fetchone()["nome"]

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requisições"
    sheet.cell(row=3, column=6, value=atividade_nome)
    sheet.cell(row=3, column=7, value=4)
    sheet.cell(row=3, column=8, value="2026-05-24")
    sheet.cell(row=3, column=9, value="Pendente")
    payload = io.BytesIO()
    workbook.save(payload)
    payload.seek(0)

    response = client.post(
        "/admin/importar_requisicoes",
        data={
            "arquivo_excel": (payload, "requisicoes.xlsx"),
        },
        content_type="multipart/form-data",
        follow_redirects=False,
    )

    assert response.status_code in (302, 303)
    with main.app.app_context():
        conn = main.get_db_connection()
        req = conn.execute(
            """
            SELECT id, atividade_versao_id, regra_snapshot_json, codigo_normativo_snapshot
              FROM requisicoes
             WHERE aluno_id IS NULL AND observacao = ?
             ORDER BY id DESC
             LIMIT 1
            """,
            ("Importado da planilha linha 3",),
        ).fetchone()
    assert req is not None
    assert req["atividade_versao_id"] is None
    assert req["regra_snapshot_json"] is None
    assert req["codigo_normativo_snapshot"] is None


def test_shadow_read_resolver_exception_captures_traceback_details(
    shadow_read_env, monkeypatch, tmp_path
):
    env = shadow_read_env
    client = env["client"]
    seed = _seed_student_in_turma()
    _set_aluno_session(client, user_id=seed["usuario_id"], user_name=seed["nome"])
    monkeypatch.setenv("SGAA_VERSIONED_RESOLVER_SHADOW_READ", "1")

    probe_log = tmp_path / "logs" / "versioned_shadow_reads.log"
    probe_log.parent.mkdir(parents=True, exist_ok=True)
    monkeypatch.setattr(
        versioning_shadow_reads,
        "_versioned_shadow_read_dedicated_log_path",
        lambda: str(probe_log),
    )

    unique_marker = uuid.uuid4().hex[:10]
    raised_message = f"boom trace {unique_marker} key=value"

    def raising_resolver(conn, *, aluno_id, atividade_id_legacy, strict_legacy_scope=True):
        raise RuntimeError(raised_message)

    monkeypatch.setattr(
        versioning_shadow_reads,
        "resolver_service",
        SimpleNamespace(resolver_versao_por_aluno=raising_resolver),
    )

    event_name = f"Evento Shadow Trace {unique_marker}"
    response = client.post(
        "/aluno/nova-requisicao",
        data={
            "atividade_id": "1",
            "nome_evento": event_name,
            "data_evento": "2026-05-24",
            "horas_solicitadas": "2",
            "observacao": "captura de traceback shadow read",
        },
        follow_redirects=False,
    )

    assert response.status_code in (302, 303)
    assert probe_log.exists()
    log_lines = probe_log.read_text(encoding="utf-8").splitlines()
    matching_lines = [
        line
        for line in log_lines
        if "event=versioned_resolver_shadow_read" in line
        and "origin=aluno_create" in line
        and "reason=resolver_exception" in line
    ]
    assert matching_lines, "Esperado evento resolver_exception no log dedicado."
    last_line = matching_lines[-1]

    assert "timestamp=" in last_line
    assert "exception_type=RuntimeError" in last_line
    assert "exception_message_b64=" in last_line
    assert "exception_traceback_b64=" in last_line

    parsed = main._parse_versioned_shadow_read_event_line(last_line)
    assert parsed is not None
    assert parsed["origin"] == "aluno_create"
    assert parsed["status"] == "error"
    assert parsed["reason"] == "resolver_exception"
    assert parsed["aluno_id"] == seed["aluno_id"]
    assert parsed["atividade_id_legacy"] == 1
    assert parsed["timestamp"] is not None
    assert "T" in parsed["timestamp"]
    assert parsed["exception_type"] == "RuntimeError"
    assert parsed["exception_message"] == raised_message
    assert parsed["exception_traceback"] is not None
    assert "RuntimeError" in parsed["exception_traceback"]
    assert raised_message in parsed["exception_traceback"]
    assert "raising_resolver" in parsed["exception_traceback"]


def test_shadow_read_resolved_event_includes_timestamp(
    shadow_read_env, monkeypatch, tmp_path
):
    env = shadow_read_env
    client = env["client"]
    seed = _seed_student_in_turma()
    _set_aluno_session(client, user_id=seed["usuario_id"], user_name=seed["nome"])
    monkeypatch.setenv("SGAA_VERSIONED_RESOLVER_SHADOW_READ", "1")

    probe_log = tmp_path / "logs" / "versioned_shadow_reads.log"
    probe_log.parent.mkdir(parents=True, exist_ok=True)
    monkeypatch.setattr(
        versioning_shadow_reads,
        "_versioned_shadow_read_dedicated_log_path",
        lambda: str(probe_log),
    )

    def fake_resolver(conn, *, aluno_id, atividade_id_legacy, strict_legacy_scope=True):
        return {
            "status": "resolved",
            "atividade_versao_id": 29,
            "atividade_base_id": 1,
            "codigo_normativo": "AAC-rev6",
            "eixo": "AAC",
            "matriz_id_efetiva": 2,
            "legacy_scope_ok": True,
            "warnings": [],
            "reason": "ok",
        }

    monkeypatch.setattr(versioning_resolver, "resolver_versao_por_aluno", fake_resolver)

    event_name = f"Evento Shadow Resolved Ts {uuid.uuid4().hex[:6]}"
    response = client.post(
        "/aluno/nova-requisicao",
        data={
            "atividade_id": "1",
            "nome_evento": event_name,
            "data_evento": "2026-05-24",
            "horas_solicitadas": "1",
            "observacao": "timestamp em resolved",
        },
        follow_redirects=False,
    )

    assert response.status_code in (302, 303)
    assert probe_log.exists()
    log_lines = probe_log.read_text(encoding="utf-8").splitlines()
    resolved_lines = [
        line
        for line in log_lines
        if "event=versioned_resolver_shadow_read" in line and "status=resolved" in line
    ]
    assert resolved_lines, "Esperado pelo menos um evento resolved no log dedicado."
    last_line = resolved_lines[-1]

    assert "timestamp=" in last_line
    assert "exception_type=" not in last_line
    assert "exception_message_b64=" not in last_line
    assert "exception_traceback_b64=" not in last_line

    parsed = main._parse_versioned_shadow_read_event_line(last_line)
    assert parsed is not None
    assert parsed["status"] == "resolved"
    assert parsed["timestamp"] is not None
    assert parsed["exception_type"] is None
    assert parsed["exception_message"] is None
    assert parsed["exception_traceback"] is None


def test_shadow_read_build_event_line_roundtrip_special_chars():
    nasty_message = 'erro foo=bar com "aspas" e quebra\nde linha'
    nasty_traceback = (
        "Traceback (most recent call last):\n"
        '  File "x.py", line 1, in foo\n'
        '    raise RuntimeError("erro foo=bar")\n'
        "RuntimeError: erro foo=bar com aspas\n"
    )
    line = main._build_versioned_shadow_read_event_line(
        origin="aluno_create",
        req_id=999,
        aluno_id=123,
        atividade_id_legacy=7,
        status="error",
        atividade_versao_id=None,
        codigo_normativo=None,
        eixo=None,
        warnings=[],
        reason="resolver_exception",
        timestamp="2026-05-27T22:30:00.123456",
        exception_type="RuntimeError",
        exception_message=nasty_message,
        exception_traceback=nasty_traceback,
    )
    # A linha gravada precisa ser monolítica para não quebrar o parser.
    assert "\n" not in line
    # Conteúdo sensível só pode vir base64 — nada cru no log line-based.
    assert nasty_message not in line
    assert nasty_traceback not in line

    parsed = main._parse_versioned_shadow_read_event_line(line)
    assert parsed is not None
    assert parsed["origin"] == "aluno_create"
    assert parsed["req_id"] == 999
    assert parsed["aluno_id"] == 123
    assert parsed["atividade_id_legacy"] == 7
    assert parsed["status"] == "error"
    assert parsed["reason"] == "resolver_exception"
    assert parsed["timestamp"] == "2026-05-27T22:30:00.123456"
    assert parsed["exception_type"] == "RuntimeError"
    assert parsed["exception_message"] == nasty_message
    assert parsed["exception_traceback"] == nasty_traceback

    # Confirma que os campos b64 são realmente base64 válido.
    payload = line.split("exception_message_b64=", 1)[1]
    msg_b64 = payload.split(" ", 1)[0]
    assert base64.b64decode(msg_b64.encode("ascii")).decode("utf-8") == nasty_message


def test_shadow_read_not_invoked_for_dashboard_and_admin_list(shadow_read_env, monkeypatch):
    env = shadow_read_env
    client = env["client"]
    seed = _seed_student_in_turma()
    monkeypatch.setenv("SGAA_VERSIONED_RESOLVER_SHADOW_READ", "1")

    calls = []

    def forbidden_resolver(conn, *, aluno_id, atividade_id_legacy, strict_legacy_scope=True):
        calls.append((aluno_id, atividade_id_legacy, strict_legacy_scope))
        raise AssertionError("resolvedor não deveria ser chamado fora da criação")

    monkeypatch.setattr(versioning_resolver, "resolver_versao_por_aluno", forbidden_resolver)

    _set_aluno_session(client, user_id=seed["usuario_id"], user_name=seed["nome"])
    aluno_dashboard = client.get("/aluno/dashboard")
    assert aluno_dashboard.status_code == 200

    _set_admin_session(client)
    admin_requisicoes = client.get("/admin/requisicoes")
    assert admin_requisicoes.status_code == 200

    assert calls == []


def test_pytest_does_not_write_to_real_general_workspace_logs():
    # Regressao D6.3.6-R / D6 dedicated log: a suite nao pode tocar logs
    # gerais reais do workspace. O log dedicado de shadow read pode existir
    # como evidencia operacional, mas nao deve ser escrito pelo pytest.
    # A fixture autouse _isolate_real_log_writes (conftest) deve estar ativa.
    real_logs_dir = os.path.abspath(os.path.join(BASE, "logs"))
    real_dedicated = os.path.join(real_logs_dir, "versioned_shadow_reads.log")
    real_app_log = os.path.join(real_logs_dir, "app.log")

    # 1) O log dedicado do shadow read aponta para fora do diretorio real.
    active_dedicated = os.path.abspath(
        versioning_shadow_reads._versioned_shadow_read_dedicated_log_path()
    )
    assert active_dedicated != os.path.abspath(real_dedicated)
    assert not active_dedicated.startswith(real_logs_dir + os.sep)

    # 2) Nenhum handler de arquivo aponta para app.log (ou app.log.1) real.
    forbidden = {
        os.path.abspath(real_app_log),
        os.path.abspath(real_app_log + ".1"),
    }
    checked_loggers = [logging.getLogger(), main.logger]
    flask_app = getattr(main, "app", None)
    if flask_app is not None:
        checked_loggers.append(flask_app.logger)
    for lg in checked_loggers:
        for handler in lg.handlers:
            base = getattr(handler, "baseFilename", None)
            if base:
                assert os.path.abspath(base) not in forbidden

    # 3) Escrever um evento de shadow read agora nao pode tocar o log dedicado
    # real do workspace; se ele existir como evidencia D6, deve permanecer
    # inalterado.
    dedicated_exists_before = os.path.exists(real_dedicated)
    dedicated_size_before = (
        os.path.getsize(real_dedicated) if dedicated_exists_before else None
    )
    main._append_versioned_shadow_read_event_line(
        "event=versioned_resolver_shadow_read origin=aluno_create req_id=1 "
        "aluno_id=1 atividade_id_legacy=1 status=resolved atividade_versao_id=2 "
        "codigo_normativo=AAC-rev6 eixo=AAC warnings=[] reason=ok"
    )
    assert os.path.exists(real_dedicated) is dedicated_exists_before
    if dedicated_exists_before:
        assert os.path.getsize(real_dedicated) == dedicated_size_before
