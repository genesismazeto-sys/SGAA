"""FC10 canonical request snapshot creation authority."""
from __future__ import annotations

import json
import uuid
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

import main
import app.views.aluno as aluno_views
from app.versioning.snapshots import RequisicaoSnapshotError, prepare_versioned_requisicao_snapshot
from tests.canonical_request_test_support import login_admin
from tests.versioned_test_support import isolated_versioned_app_env


@pytest.fixture()
def fc10_env(tmp_path):
    with isolated_versioned_app_env(tmp_path, "fc10.db") as env:
        yield env


def _student(conn):
    return conn.execute(
        "SELECT a.id AS aluno_id,a.usuario_id FROM alunos a WHERE a.matricula='PPA.TESTE.0001'"
    ).fetchone()


def _login_student(client, row):
    with client.session_transaction() as session:
        session.update(user_id=row["usuario_id"], user_type="aluno", user_name="Aluno")


def _post(client, version_id, name):
    return client.post("/aluno/nova-requisicao", data={
        "atividade_versao_id": "" if version_id is None else str(version_id),
        "nome_evento": name, "data_evento": "2026-05-10", "horas_solicitadas": "4",
    })


def _request(conn, name):
    return conn.execute("SELECT * FROM requisicoes WHERE nome_evento=?", (name,)).fetchone()


def test_student_creation_persists_exact_matrix_snapshot(fc10_env):
    with main.app.app_context():
        row = _student(main.get_db_connection())
    _login_student(fc10_env["client"], row)
    name = f"FC10-{uuid.uuid4().hex}"
    assert _post(fc10_env["client"], 29, name).status_code == 302
    with main.app.app_context():
        saved = _request(main.get_db_connection(), name)
        payload = json.loads(saved["regra_snapshot_json"])
    assert saved["atividade_versao_id"] == 29
    assert payload["atividade_versao_id"] == 29
    assert payload["atividade_versao_numero"] >= 1
    assert payload["schema_version"] == "prod-1-request-v2"
    assert {"norma_id", "codigo_normativo"}.isdisjoint(payload)


@pytest.mark.parametrize("version_id", [None, 1, 999999])
def test_missing_or_outside_exact_version_is_transactionally_rejected(fc10_env, version_id):
    with main.app.app_context():
        row = _student(main.get_db_connection())
    _login_student(fc10_env["client"], row)
    name = f"FC10-reject-{version_id}-{uuid.uuid4().hex}"
    assert _post(fc10_env["client"], version_id, name).status_code == 200
    with main.app.app_context():
        assert _request(main.get_db_connection(), name) is None


def test_missing_matrix_link_and_inactive_version_fail_closed(fc10_env):
    with main.app.app_context():
        conn = main.get_db_connection()
        student = _student(conn)
        conn.execute("DELETE FROM matriz_atividade_versao_item WHERE matriz_id=2 AND atividade_versao_id=29")
        with pytest.raises(RequisicaoSnapshotError):
            prepare_versioned_requisicao_snapshot(
                conn, flow_origin="student_create", aluno_id=student["aluno_id"], atividade_versao_id=29
            )
        conn.rollback()
        conn.execute("UPDATE atividade_versao SET status='inativa' WHERE id=29")
        with pytest.raises(RequisicaoSnapshotError):
            prepare_versioned_requisicao_snapshot(
                conn, flow_origin="student_create", aluno_id=student["aluno_id"], atividade_versao_id=29
            )


def test_matrix_db_constraint_prevents_ambiguous_version_for_same_base(fc10_env):
    with main.app.app_context():
        conn = main.get_db_connection()
        base_id = conn.execute("SELECT atividade_base_id FROM atividade_versao WHERE id=29").fetchone()[0]
        other = conn.execute(
            "SELECT id FROM atividade_versao WHERE atividade_base_id=? AND id<>29 LIMIT 1", (base_id,)
        ).fetchone()[0]
        with pytest.raises(Exception):
            conn.execute(
                "INSERT INTO matriz_atividade_versao_item(matriz_id,atividade_base_id,atividade_versao_id) VALUES(2,?,?)",
                (base_id, other),
            )


def test_attachment_failure_rolls_back_request_and_compensates_saved_file(fc10_env, monkeypatch):
    with main.app.app_context():
        student = _student(main.get_db_connection())
    _login_student(fc10_env["client"], student)
    saved_paths = []

    def failing_second_save(_upload, _allowed, *, root_folder, **_kwargs):
        if saved_paths:
            raise OSError("second attachment failed")
        relative = "r3/first-proof.pdf"
        absolute = Path(root_folder) / relative
        absolute.parent.mkdir(parents=True, exist_ok=True)
        absolute.write_bytes(b"proof")
        saved_paths.append((relative, absolute))
        return relative

    monkeypatch.setattr(aluno_views, "save_student_document", failing_second_save)
    name = f"FC10-compensation-{uuid.uuid4().hex}"
    response = fc10_env["client"].post(
        "/aluno/nova-requisicao",
        data={
            "atividade_versao_id": "29",
            "nome_evento": name,
            "data_evento": "2026-05-10",
            "horas_solicitadas": "4",
            "comprovantes_files": [
                (BytesIO(b"one"), "one.pdf"),
                (BytesIO(b"two"), "two.pdf"),
            ],
            "comprovantes_labels": ["one", "two"],
        },
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    with main.app.app_context():
        assert _request(main.get_db_connection(), name) is None
    assert saved_paths and not saved_paths[0][1].exists()


def test_admin_import_producer_persists_mandatory_exact_snapshot(fc10_env):
    login_admin(fc10_env["client"])
    with main.app.app_context():
        conn = main.get_db_connection()
        activity_name = conn.execute(
            """SELECT b.nome_conceito FROM atividade_versao v
                 JOIN atividade_base b ON b.id=v.atividade_base_id WHERE v.id=29"""
        ).fetchone()[0]

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requisições"
    sheet.cell(row=3, column=1, value="PPA.TESTE.0001")
    sheet.cell(row=3, column=6, value=activity_name)
    sheet.cell(row=3, column=7, value=4)
    sheet.cell(row=3, column=8, value="2026-05-10")
    sheet.cell(row=3, column=9, value="Deferido")
    payload = BytesIO()
    workbook.save(payload)
    payload.seek(0)

    response = fc10_env["client"].post(
        "/admin/importar_requisicoes",
        data={"arquivo_excel": (payload, "fc10-import.xlsx")},
        content_type="multipart/form-data",
    )
    assert response.status_code in (302, 303)
    with main.app.app_context():
        rows = main.get_db_connection().execute(
            "SELECT * FROM requisicoes WHERE observacao LIKE 'Importado da planilha linha 3%'"
        ).fetchall()
        assert len(rows) == 1
        saved = rows[0]
        snapshot = json.loads(saved["regra_snapshot_json"])
        assert saved["atividade_versao_id"] == 29
        assert "codigo_normativo_snapshot" not in saved.keys()
        assert {"norma_id", "codigo_normativo"}.isdisjoint(snapshot)
        assert snapshot["atividade_versao_id"] == 29
        assert snapshot["flow_origin"] == "admin_import"
