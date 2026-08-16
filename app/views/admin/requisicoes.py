from __future__ import annotations

import logging
import os
import shutil
import traceback
import datetime

import openpyxl

from flask import Blueprint, current_app, jsonify, redirect, render_template, request, session, url_for

from app.activity_catalog import parse_documentos_json
from app.auth import admin_required
from app.db import ensure_turmas_matriz_schema, get_db_connection
from app.db_maintenance import ensure_matriz_atividade_links_table
from app.matrix_scope import (
    _matriz_option_label,
    get_allowed_activity_ids_for_turma_matrix,
    is_activity_allowed_for_turma_matrix,
)
from app.requisitions import auto_indefer_devolvidas
from app.student_documents import remove_student_document, save_student_document
from app.text import normalize_header
from app.uploads import ALLOWED_ATTACHMENTS, _allowed, save_upload
from app.versioning.shadow_reads import maybe_run_versioned_resolver_shadow_read
from app.versioning.snapshots import (
    _build_admin_requisicao_snapshot_diagnostic,
    _has_versioned_requisicao_snapshot,
    is_versioned_requisicao_snapshot_display_enabled,
    prepare_versioned_requisicao_snapshot,
    read_requisicao_snapshot_for_processing,
    RequisicaoSnapshotError,
    SnapshotProcessingAuthority,
)
from app.web.filters import append_conditions_sql, get_date_range_query, get_multi_query_values
from app.web.pagination import get_pagination, wants_pagination
from utils.messages import flash

from app.views.admin import LegacyRouteSpec, configure_legacy_routes


logger = logging.getLogger("main")
logger.setLevel(logging.INFO)


ALLOWED_EXCEL = {"xlsx"}


@admin_required
def admin_importar_requisicoes():
    if request.method == "POST":
        usar_padrao = "usar_arquivo_padrao" in request.form
        arquivo_selecionado = request.files.get("arquivo_excel")

        arquivo_path = None
        if usar_padrao:
            arquivo_path = os.path.join(current_app.config["UPLOAD_FOLDER"], "Acompanhamento de atividades complementares.xlsx")
            if not os.path.exists(arquivo_path):
                flash("Arquivo padrão não encontrado na pasta de uploads.", "error")
                return render_template("admin_importar_requisicoes.html")
        elif arquivo_selecionado and arquivo_selecionado.filename != "":
            if not _allowed(arquivo_selecionado.filename, ALLOWED_EXCEL):
                flash("Envie um arquivo .xlsx válido.", "error")
                return render_template("admin_importar_requisicoes.html")
            filename = save_upload(arquivo_selecionado, ALLOWED_EXCEL, prefix="import", subdir="imports")
            arquivo_path = os.path.join(current_app.config["UPLOAD_FOLDER"], filename)
        else:
            flash("Nenhum arquivo selecionado.", "error")
            return render_template("admin_importar_requisicoes.html")

        try:
            logger.info(f"Iniciando importação do arquivo: {arquivo_path} usando openpyxl")
            workbook = openpyxl.load_workbook(arquivo_path, data_only=True)
            if "Requisições" not in workbook.sheetnames:
                flash("Aba 'Requisições' não encontrada na planilha.", "error")
                logger.error(f"Aba 'Requisições' não encontrada em {arquivo_path}")
                return render_template("admin_importar_requisicoes.html")

            sheet = workbook["Requisições"]
            logger.info(f"Planilha 'Requisições' lida. Total de linhas: {sheet.max_row}")

            conn = get_db_connection()
            sucesso_count = 0
            erro_count = 0
            erros_detalhes = []

            atividades_map = {normalize_header(row["nome"]): row["id"] for row in conn.execute("SELECT id, nome FROM atividades").fetchall()}
            logger.info(f"Cache de atividades criado: {len(atividades_map)} atividades.")

            data_solicitacao_hoje = datetime.date.today().strftime("%Y-%m-%d")

            for row_index in range(3, sheet.max_row + 1):
                try:
                    nome_atividade_raw = sheet.cell(row=row_index, column=6).value
                    data_evento_raw = sheet.cell(row=row_index, column=8).value
                    horas_raw = sheet.cell(row=row_index, column=7).value
                    status_raw = sheet.cell(row=row_index, column=9).value

                    if nome_atividade_raw is None or data_evento_raw is None or horas_raw is None:
                        logger.warning(f"Linha {row_index}: dados essenciais ausentes, pulando.")
                        continue
                    if all(sheet.cell(row=row_index, column=c).value is None for c in range(1, sheet.max_column + 1)):
                        logger.info(f"Linha {row_index}: linha vazia, pulando.")
                        continue

                    aluno_id = None
                    nome_atividade_norm = normalize_header(str(nome_atividade_raw))
                    atividade_id = atividades_map.get(nome_atividade_norm)
                    if not atividade_id:
                        logger.warning(f"Linha {row_index}: Atividade '{nome_atividade_raw}' não encontrada")
                        erro_count += 1
                        erros_detalhes.append(f"Linha {row_index}: Atividade '{nome_atividade_raw}' não encontrada")
                        continue

                    data_evento = "Indisponível"
                    if isinstance(data_evento_raw, datetime.datetime):
                        data_evento = data_evento_raw.strftime("%Y-%m-%d")
                    elif data_evento_raw is not None:
                        try:
                            if isinstance(data_evento_raw, (int, float)):
                                data_evento_dt = openpyxl.utils.datetime.from_excel(data_evento_raw)
                                data_evento = data_evento_dt.strftime("%Y-%m-%d")
                            else:
                                data_evento_dt = datetime.datetime.strptime(str(data_evento_raw).split()[0], '%Y-%m-%d')
                                data_evento = data_evento_dt.strftime("%Y-%m-%d")
                        except (ValueError, TypeError) as e_date:
                            logger.warning(f"Linha {row_index}: data inválida '{data_evento_raw}'. {e_date}")

                    horas_solicitadas = 0.0
                    horas_deferidas = None
                    if horas_raw is not None:
                        try:
                            horas_solicitadas = float(horas_raw)
                        except (ValueError, TypeError) as e_horas:
                            logger.warning(f"Linha {row_index}: horas inválidas '{horas_raw}'. {e_horas}")
                    else:
                        logger.warning(f"Linha {row_index}: horas ausentes, usando 0.0.")

                    status = "Pendente"
                    if status_raw is not None:
                        status_norm = normalize_header(str(status_raw))
                        if status_norm == "deferido":
                            status = "Deferida"
                            horas_deferidas = horas_solicitadas
                        elif status_norm == "deferido parcialmente":
                            status = "Deferida Parcialmente"
                            horas_deferidas = 0.0
                        elif status_norm == "indeferido":
                            status = "Indeferida"
                            horas_deferidas = 0.0

                    conn.execute("""
                        INSERT INTO requisicoes\x20
                        (aluno_id, atividade_id, data_solicitacao, data_evento, horas_solicitadas, status, horas_deferidas, observacao)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (aluno_id, atividade_id, data_solicitacao_hoje, data_evento, horas_solicitadas, status, horas_deferidas, f"Importado da planilha linha {row_index}"))
                    sucesso_count += 1

                except Exception as e:
                    logger.error(f"Erro ao processar linha {row_index}: {e}")
                    traceback.print_exc()
                    erro_count += 1
                    erros_detalhes.append(f"Linha {row_index}: Erro inesperado - {e}")

            conn.commit()
            logger.info(f"Importação concluída. Sucesso: {sucesso_count}, Erros/Pulados: {erro_count}")

            flash(f"{sucesso_count} requisições importadas com sucesso.", "success")
            if erro_count > 0:
                flash(f"{erro_count} linhas não puderam ser importadas ou foram puladas. Veja app.log para detalhes.", "warning")

            return redirect(url_for("admin_requisicoes"))

        except FileNotFoundError:
            logger.error(f"Arquivo não encontrado: {arquivo_path}")
            flash(f"Erro: Arquivo não encontrado em {arquivo_path}", "error")
            return render_template("admin_importar_requisicoes.html")
        except Exception as e:
            logger.error(f"Erro GERAL durante a importação: {e}")
            traceback.print_exc()
            flash(f"Ocorreu um erro grave durante a importação: {e}", "error")
            return render_template("admin_importar_requisicoes.html")

    return render_template("admin_importar_requisicoes.html")


def _normalize_requisicao_data_evento(value):
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    if "/" in raw:
        parts = raw.split("/")
        if len(parts) >= 3:
            try:
                return f"{parts[2]}-{int(parts[1]):02d}-{int(parts[0]):02d}"
            except Exception:
                return None
        return None
    if "-" in raw:
        return raw.split(" ")[0][:10]
    return None


def _get_admin_requisicao_scope_for_aluno(conn, aluno_id):
    ensure_turmas_matriz_schema(conn)
    ensure_matriz_atividade_links_table(conn)
    row = conn.execute(
        """
        SELECT a.id, a.nome, a.matricula, a.turma_id,
               t.nome AS turma_nome, t.codigo AS turma_codigo,
               t.curso_id, t.matriz_id AS turma_matriz_id
          FROM alunos a
          LEFT JOIN turmas t ON t.id = a.turma_id
         WHERE a.id = ?
        """,
        (aluno_id,),
    ).fetchone()
    if not row:
        return None
    allowed_activity_ids, matriz = get_allowed_activity_ids_for_turma_matrix(
        conn,
        row["curso_id"],
        row["turma_matriz_id"],
    )
    turma_label = row["turma_codigo"] or row["turma_nome"] or "Sem turma"
    return {
        "aluno": row,
        "allowed_activity_ids": sorted(allowed_activity_ids) if allowed_activity_ids is not None else None,
        "matriz_scope": ({"id": matriz["id"], "label": _matriz_option_label(matriz)} if matriz else None),
        "turma_label": turma_label,
    }


def _list_admin_requisicao_alunos(conn):
    ensure_turmas_matriz_schema(conn)
    return conn.execute(
        """
        SELECT a.id, a.nome, a.matricula,
               COALESCE(t.codigo, t.nome, 'Sem turma') AS turma_label
          FROM alunos a
          LEFT JOIN turmas t ON t.id = a.turma_id
         ORDER BY a.nome COLLATE NOCASE, a.id
        """
    ).fetchall()


def _append_requisicao_arquivos(
    conn,
    req_id,
    aluno_id,
    arquivos,
    labels=None,
    created_document_paths=None,
):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS requisicao_arquivos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            requisicao_id INTEGER NOT NULL,
            label TEXT,
            filename TEXT,
            criado_em TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(requisicao_id) REFERENCES requisicoes(id)
        )
        """
    )
    first_saved = None
    labels = labels or []
    aluno_row = conn.execute("SELECT nome FROM alunos WHERE id = ?", (aluno_id,)).fetchone()
    student_name = str((aluno_row["nome"] if aluno_row else "") or f"aluno-{aluno_id}")
    for idx, arquivo in enumerate(arquivos or []):
        if not arquivo or not getattr(arquivo, "filename", ""):
            continue
        if not _allowed(arquivo.filename, ALLOWED_ATTACHMENTS):
            flash(f"Arquivo ignorado por extensão não permitida: {arquivo.filename}", "warning")
            continue
        saved = None
        try:
            saved = save_student_document(
                arquivo,
                ALLOWED_ATTACHMENTS,
                root_folder=current_app.config["DOCUMENTOS_ALUNOS_FOLDER"],
                student_id=aluno_id,
                student_name=student_name,
                category="requisicoes",
                prefix=f"req{req_id}",
            )
            if saved:
                if created_document_paths is not None:
                    created_document_paths.append(saved)
                label_value = labels[idx] if labels and idx < len(labels) else None
                conn.execute(
                    "INSERT INTO requisicao_arquivos (requisicao_id, label, filename) VALUES (?, ?, ?)",
                    (req_id, label_value, saved),
                )
                if first_saved is None:
                    first_saved = saved
        except Exception as exc:
            if created_document_paths is None and saved:
                try:
                    remove_student_document(
                        current_app.config["DOCUMENTOS_ALUNOS_FOLDER"], saved
                    )
                except Exception:
                    logger.exception("Falha ao compensar arquivo de comprovante")
            if created_document_paths is not None:
                raise
            logger.error(f"Falha ao salvar arquivo de comprovante na requisição {req_id}: {exc}")
    return first_saved


@admin_required
def admin_requisicoes():
    page, per_page, offset = get_pagination(default_per_page=25)
    status_filters = {item.strip().lower() for item in get_multi_query_values("status") if item.strip()}
    processamento_filters = {
        item.strip().lower()
        for item in get_multi_query_values("processamento")
        if item.strip().lower() in {"com_data", "sem_data"}
    }
    aluno_filters = [item for item in get_multi_query_values("aluno") if item]
    turma_filters = [item for item in get_multi_query_values("turma") if item]
    tipo_filters = [item for item in get_multi_query_values("tipo") if item]
    grupo_filters = [item for item in get_multi_query_values("grupo") if item]
    atividade_filters = [item for item in get_multi_query_values("atividade") if item]
    data_solicitacao_min, data_solicitacao_max = get_date_range_query("data_solicitacao")
    data_processamento_min, data_processamento_max = get_date_range_query("data_processamento")
    status_filtro = next(iter(status_filters), 'Todas') if status_filters else 'Todas'
    q = (request.args.get('q') or '').strip()
    sort_field = (request.args.get('s') or 'data_solicitacao').strip()
    sort_dir = (request.args.get('dir') or 'desc').strip().lower()
    conn = get_db_connection()
    auto_indefer_devolvidas(conn)
    ensure_turmas_matriz_schema(conn)
    ensure_matriz_atividade_links_table(conn)
    base_from = """
        FROM requisicoes r
        LEFT JOIN alunos a ON r.aluno_id = a.id
        LEFT JOIN turmas t ON t.id = a.turma_id
        JOIN atividades act ON r.atividade_id = act.id
    """
    select_cols = """
        SELECT r.*,\x20
               a.nome              AS aluno_nome,
               a.matricula         AS aluno_matricula,
               COALESCE(t.codigo, t.nome, a.turma) AS turma_codigo,
               t.curso_id          AS turma_curso_id,
               t.matriz_id         AS turma_matriz_id,
               act.nome            AS atividade_nome,
               act.grupo           AS grupo,
               act.tipo_atividade  AS tipo_atividade
    """
    query = select_cols + base_from
    params = []
    where = []
    if aluno_filters:
        placeholders = ", ".join("?" for _ in aluno_filters)
        where.append(f"COALESCE(TRIM(a.nome), '') IN ({placeholders})")
        params.extend(aluno_filters)
    if turma_filters:
        placeholders = ", ".join("?" for _ in turma_filters)
        where.append(f"COALESCE(TRIM(COALESCE(t.codigo, t.nome, a.turma)), '') IN ({placeholders})")
        params.extend(turma_filters)
    if tipo_filters:
        placeholders = ", ".join("?" for _ in tipo_filters)
        where.append(f"COALESCE(TRIM(act.tipo_atividade), '') IN ({placeholders})")
        params.extend(tipo_filters)
    if grupo_filters:
        placeholders = ", ".join("?" for _ in grupo_filters)
        where.append(f"COALESCE(TRIM(act.grupo), '') IN ({placeholders})")
        params.extend(grupo_filters)
    if atividade_filters:
        placeholders = ", ".join("?" for _ in atividade_filters)
        where.append(f"COALESCE(TRIM(act.nome), '') IN ({placeholders})")
        params.extend(atividade_filters)
    if data_solicitacao_min:
        where.append("date(r.data_solicitacao) >= date(?)")
        params.append(data_solicitacao_min)
    if data_solicitacao_max:
        where.append("date(r.data_solicitacao) <= date(?)")
        params.append(data_solicitacao_max)
    if data_processamento_min:
        where.append("date(r.data_processamento) >= date(?)")
        params.append(data_processamento_min)
    if data_processamento_max:
        where.append("date(r.data_processamento) <= date(?)")
        params.append(data_processamento_max)
    if processamento_filters:
        processamento_clauses = []
        if "com_data" in processamento_filters:
            processamento_clauses.append("(r.data_processamento IS NOT NULL AND TRIM(r.data_processamento) <> '')")
        if "sem_data" in processamento_filters:
            processamento_clauses.append("(r.data_processamento IS NULL OR TRIM(r.data_processamento) = '')")
        if processamento_clauses:
            where.append("(" + " OR ".join(processamento_clauses) + ")")
    if status_filters:
        status_clauses = []
        for status_filter in status_filters:
            if status_filter == 'pendente':
                status_clauses.append("LOWER(COALESCE(r.status, '')) IN ('pendente', 'aguardando', 'pending')")
            elif status_filter == 'deferida':
                status_clauses.append("LOWER(COALESCE(r.status, '')) IN ('deferido', 'deferida', 'aprovado', 'aprovada', 'approved')")
            elif status_filter == 'deferida_parcialmente':
                status_clauses.append("LOWER(COALESCE(r.status, '')) IN ('deferida parcialmente', 'deferido parcialmente', 'parcialmente deferida', 'partially approved')")
            elif status_filter == 'indeferida':
                status_clauses.append("LOWER(COALESCE(r.status, '')) IN ('indeferido', 'indeferida', 'rejeitado', 'rejeitada')")
            elif status_filter == 'devolvida':
                status_clauses.append("LOWER(COALESCE(r.status, '')) IN ('devolvido', 'devolvida')")
            elif status_filter == 'encerrada':
                status_clauses.append("LOWER(COALESCE(r.status, '')) IN ('encerrado', 'encerrada', 'closed')")
            else:
                status_clauses.append("LOWER(COALESCE(r.status, '')) = ?")
                params.append(status_filter)
        if status_clauses:
            where.append("(" + " OR ".join(status_clauses) + ")")
    if q:
        like = f"%{q}%"
        where.append("(a.nome LIKE ? OR a.matricula LIKE ? OR a.turma LIKE ? OR act.nome LIKE ? OR act.grupo LIKE ? OR act.tipo_atividade LIKE ? OR r.status LIKE ?)")
        params.extend([like, like, like, like, like, like, like])
    where_sql = append_conditions_sql(False, where)
    query += where_sql
    # Ordenação (whitelist para segurança)
    order_map = {
        'data_solicitacao': 'r.data_solicitacao',
        'data_processamento': 'r.data_processamento',
        'aluno_nome': 'a.nome',
        'turma_codigo': 'a.turma',
        'tipo_atividade': 'act.tipo_atividade',
        'grupo': 'act.grupo',
        'atividade_nome': 'act.nome',
        'status': 'r.status'
    }
    col = order_map.get(sort_field, 'r.data_solicitacao')
    direction = 'DESC' if sort_dir == 'desc' else 'ASC'
    query += f" ORDER BY {col} {direction}"

    # total para paginação (conta sem ORDER BY)
    count_sql = "SELECT COUNT(*) " + base_from + where_sql
    total = conn.execute(count_sql, params).fetchone()[0]

    # Só aplica LIMIT/OFFSET se o usuário explicitamente paginar (evita quebrar UI sem controles)
    apply_limit = wants_pagination()
    params_exec = list(params)
    if apply_limit:
        query += " LIMIT ? OFFSET ?"
        params_exec += [per_page, offset]

    requisicoes_rows = conn.execute(query, params_exec).fetchall()
    requisicoes = []
    matrix_scope_cache = {}
    for row in requisicoes_rows:
        item = {k: row[k] for k in row.keys()}
        item["snapshot_versionado_presente"] = _has_versioned_requisicao_snapshot(item)
        cache_key = (item.get("turma_curso_id"), item.get("turma_matriz_id"))
        if cache_key not in matrix_scope_cache:
            matrix_scope_cache[cache_key] = get_allowed_activity_ids_for_turma_matrix(
                conn,
                item.get("turma_curso_id"),
                item.get("turma_matriz_id"),
            )
        allowed_activity_ids, matriz = matrix_scope_cache[cache_key]
        item["matrix_scope_issue"] = item.get("atividade_id") not in allowed_activity_ids
        item["matrix_scope_label"] = _matriz_option_label(matriz) if matriz else None
        requisicoes.append(item)
    # Carregar atividades e documentos obrigatórios (para reuso do form do aluno no modal admin)
    atividades = conn.execute("SELECT * FROM atividades ORDER BY tipo_atividade, grupo, nome").fetchall()
    alunos_opcoes = _list_admin_requisicao_alunos(conn)
    docs_por_atividade = {}
    try:
        for a in atividades:
            raw = None
            try:
                raw = a["documentos_json"] if "documentos_json" in a.keys() else None
            except Exception:
                raw = None
            docs = parse_documentos_json(raw)
            if docs:
                docs_por_atividade[a["id"]] = docs
    except Exception:
        pass
    alunos_filtro = conn.execute(
        """
        SELECT DISTINCT COALESCE(NULLIF(TRIM(a.nome), ''), '') AS aluno_nome
          FROM requisicoes r
          LEFT JOIN alunos a ON r.aluno_id = a.id
         WHERE COALESCE(NULLIF(TRIM(a.nome), ''), '') <> ''
      ORDER BY LOWER(COALESCE(a.nome, '')) ASC
        """
    ).fetchall()
    turmas_filtro = conn.execute(
        """
        SELECT DISTINCT COALESCE(NULLIF(TRIM(COALESCE(t.codigo, t.nome, a.turma)), ''), '') AS turma_codigo
          FROM requisicoes r
          LEFT JOIN alunos a ON r.aluno_id = a.id
          LEFT JOIN turmas t ON t.id = a.turma_id
         WHERE COALESCE(NULLIF(TRIM(COALESCE(t.codigo, t.nome, a.turma)), ''), '') <> ''
      ORDER BY LOWER(COALESCE(t.codigo, t.nome, a.turma, '')) ASC
        """
    ).fetchall()
    tipos_filtro = conn.execute(
        """
        SELECT DISTINCT COALESCE(NULLIF(TRIM(act.tipo_atividade), ''), '') AS tipo_atividade
          FROM requisicoes r
          JOIN atividades act ON r.atividade_id = act.id
         WHERE COALESCE(NULLIF(TRIM(act.tipo_atividade), ''), '') <> ''
      ORDER BY LOWER(COALESCE(act.tipo_atividade, '')) ASC
        """
    ).fetchall()
    grupos_filtro = conn.execute(
        """
        SELECT DISTINCT COALESCE(NULLIF(TRIM(act.grupo), ''), '') AS grupo
          FROM requisicoes r
          JOIN atividades act ON r.atividade_id = act.id
         WHERE COALESCE(NULLIF(TRIM(act.grupo), ''), '') <> ''
      ORDER BY LOWER(COALESCE(act.grupo, '')) ASC
        """
    ).fetchall()
    atividades_filtro = conn.execute(
        """
        SELECT DISTINCT COALESCE(NULLIF(TRIM(act.nome), ''), '') AS atividade_nome
          FROM requisicoes r
          JOIN atividades act ON r.atividade_id = act.id
         WHERE COALESCE(NULLIF(TRIM(act.nome), ''), '') <> ''
      ORDER BY LOWER(COALESCE(act.nome, '')) ASC
        """
    ).fetchall()
    filter_schema = [
        {
            "param": "data_solicitacao",
            "label": "Solicitação",
            "type": "date_range",
            "min_label": "De",
            "max_label": "Até",
        },
        {
            "param": "data_processamento",
            "label": "Processamento",
            "type": "date_range",
            "min_label": "De",
            "max_label": "Até",
        },
        {
            "param": "processamento",
            "label": "Estado do processamento",
            "type": "multi_select",
            "values": [
                {"value": "com_data", "label": "Com processamento"},
                {"value": "sem_data", "label": "Sem processamento"},
            ],
        },
        {
            "param": "aluno",
            "label": "Aluno",
            "type": "multi_select",
            "values": [
                {"value": row["aluno_nome"], "label": row["aluno_nome"]}
                for row in alunos_filtro
            ],
        },
        {
            "param": "turma",
            "label": "Turma",
            "type": "multi_select",
            "values": [
                {"value": row["turma_codigo"], "label": row["turma_codigo"]}
                for row in turmas_filtro
            ],
        },
        {
            "param": "tipo",
            "label": "Tipo",
            "type": "multi_select",
            "values": [
                {"value": row["tipo_atividade"], "label": row["tipo_atividade"]}
                for row in tipos_filtro
            ],
        },
        {
            "param": "grupo",
            "label": "Grupo",
            "type": "multi_select",
            "values": [
                {"value": row["grupo"], "label": row["grupo"]}
                for row in grupos_filtro
            ],
        },
        {
            "param": "atividade",
            "label": "Atividade",
            "type": "multi_select",
            "values": [
                {"value": row["atividade_nome"], "label": row["atividade_nome"]}
                for row in atividades_filtro
            ],
        },
        {
            "param": "status",
            "label": "Status",
            "type": "multi_select",
            "values": [
                {"value": "pendente", "label": "Pendente"},
                {"value": "deferida", "label": "Deferida"},
                {"value": "deferida_parcialmente", "label": "Deferida Parcialmente"},
                {"value": "indeferida", "label": "Indeferida"},
                {"value": "devolvida", "label": "Devolvida"},
                {"value": "encerrada", "label": "Encerrada"},
            ],
        },
    ]
    total_pages = (total + per_page - 1) // per_page if apply_limit and per_page else 1
    return render_template(
        "admin_requisicoes.html",
        requisicoes=requisicoes,
        status_atual=status_filtro,
        page=page,
        per_page=per_page,
        total=total,
        total_pages=total_pages,
        atividades=atividades,
        alunos_opcoes=alunos_opcoes,
        docs_por_atividade=docs_por_atividade,
        filter_schema=filter_schema,
    )


@admin_required
def admin_nova_requisicao():
    conn = get_db_connection()
    ensure_turmas_matriz_schema(conn)
    ensure_matriz_atividade_links_table(conn)

    if request.method == "GET":
        aluno_id = (request.args.get("aluno_id") or "").strip()
        return redirect(url_for("admin_requisicoes", open_new="1", aluno_id=aluno_id or None))

    aluno_id = request.form.get("aluno_id", type=int)
    atividade_id = request.form.get("atividade_id", type=int)
    nome_evento = (request.form.get("nome_evento") or "").strip()
    observacao = (request.form.get("observacao") or "").strip() or None
    data_evento = _normalize_requisicao_data_evento(request.form.get("data_evento"))
    horas_raw = (request.form.get("horas_solicitadas") or "").strip()

    redirect_kwargs = {"open_new": "1"}
    if aluno_id:
        redirect_kwargs["aluno_id"] = aluno_id

    scope = _get_admin_requisicao_scope_for_aluno(conn, aluno_id) if aluno_id else None
    if not scope:
        flash("Selecione um aluno válido para criar a requisição.", "error")
        return redirect(url_for("admin_requisicoes", **redirect_kwargs))

    if not atividade_id:
        flash("Selecione uma atividade válida.", "error")
        return redirect(url_for("admin_requisicoes", **redirect_kwargs))

    allowed_activity_ids = scope["allowed_activity_ids"]
    if atividade_id not in allowed_activity_ids:
        flash("A atividade selecionada não pertence à matriz efetiva da turma do aluno.", "error")
        return redirect(url_for("admin_requisicoes", **redirect_kwargs))

    atividade = conn.execute("SELECT id FROM atividades WHERE id = ?", (atividade_id,)).fetchone()
    if not atividade:
        flash("Atividade não encontrada.", "error")
        return redirect(url_for("admin_requisicoes", **redirect_kwargs))

    if not nome_evento:
        flash("Informe o nome do evento.", "error")
        return redirect(url_for("admin_requisicoes", **redirect_kwargs))

    try:
        horas_solicitadas = float(horas_raw)
        if horas_solicitadas < 0:
            raise ValueError()
    except Exception:
        flash("Informe um valor válido para horas solicitadas.", "error")
        return redirect(url_for("admin_requisicoes", **redirect_kwargs))

    if not data_evento:
        flash("Informe uma data válida para o evento.", "error")
        return redirect(url_for("admin_requisicoes", **redirect_kwargs))

    try:
        prepared_snapshot = prepare_versioned_requisicao_snapshot(
            conn,
            flow_origin="admin_create",
            aluno_id=aluno_id,
            atividade_id_legacy=atividade_id,
        )
    except RequisicaoSnapshotError as exc:
        conn.rollback()
        flash(exc.user_message, "error")
        return redirect(url_for("admin_requisicoes", **redirect_kwargs))
    except Exception:
        conn.rollback()
        logger.exception("Falha ao preparar snapshot obrigatório da requisição do admin")
        raise

    data_solicitacao = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    arquivos = request.files.getlist("comprovantes_files") or []
    created_document_paths = []
    try:
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO requisicoes
            (aluno_id, atividade_id, data_solicitacao, data_evento, horas_solicitadas, nome_evento, status, observacao, arquivo_comprovante,
             atividade_versao_id, regra_snapshot_json, codigo_normativo_snapshot)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                aluno_id,
                atividade_id,
                data_solicitacao,
                data_evento,
                horas_solicitadas,
                nome_evento,
                "Pendente",
                observacao,
                None,
                prepared_snapshot.atividade_versao_id,
                prepared_snapshot.snapshot_json,
                prepared_snapshot.codigo_normativo,
            ),
        )
        req_id = cur.lastrowid
        first_saved = _append_requisicao_arquivos(
            conn,
            req_id,
            aluno_id,
            arquivos,
            created_document_paths=created_document_paths,
        )

        if first_saved:
            conn.execute("UPDATE requisicoes SET arquivo_comprovante = ? WHERE id = ?", (first_saved, req_id))

        conn.commit()
        created_document_paths.clear()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            logger.exception("Falha ao reverter requisição do admin")
        for rel_path in created_document_paths:
            try:
                remove_student_document(
                    current_app.config["DOCUMENTOS_ALUNOS_FOLDER"], rel_path
                )
            except Exception:
                    logger.exception("Falha ao compensar arquivo de comprovante")
        logger.exception("Falha ao criar requisição do admin")
        return redirect(url_for("admin_requisicoes", **redirect_kwargs))
    try:
        maybe_run_versioned_resolver_shadow_read(
            conn,
            origin="admin_create",
            aluno_id=aluno_id,
            atividade_id_legacy=atividade_id,
            req_id=req_id,
        )
    except Exception:
        logger.exception("Falha ao executar resolvedor versionado em modo sombra no fluxo do admin")
    flash("Requisição criada com sucesso.", "success")
    return redirect(url_for("admin_requisicoes"))


@admin_required
def admin_editar_requisicao(req_id):
    conn = get_db_connection()
    ensure_turmas_matriz_schema(conn)
    ensure_matriz_atividade_links_table(conn)

    requisicao = conn.execute(
        """
        SELECT r.*, a.turma_id,
               t.curso_id AS turma_curso_id,
               t.matriz_id AS turma_matriz_id
          FROM requisicoes r
          LEFT JOIN alunos a ON a.id = r.aluno_id
          LEFT JOIN turmas t ON t.id = a.turma_id
         WHERE r.id = ?
        """,
        (req_id,),
    ).fetchone()
    if not requisicao:
        flash("Requisição não encontrada.", "error")
        return redirect(url_for("admin_requisicoes"))

    redirect_kwargs = {"open_edit": "1", "req_id": req_id}

    if (requisicao["status"] or "") != "Pendente":
        flash("Somente requisições pendentes podem ser editadas pelo admin.", "error")
        return redirect(url_for("admin_requisicoes", **redirect_kwargs))

    atividade_id = request.form.get("atividade_id", type=int)
    nome_evento = (request.form.get("nome_evento") or "").strip()
    observacao = (request.form.get("observacao") or "").strip() or None
    data_evento = _normalize_requisicao_data_evento(request.form.get("data_evento"))
    horas_raw = (request.form.get("horas_solicitadas") or "").strip()

    if not nome_evento:
        flash("Informe o nome do evento.", "error")
        return redirect(url_for("admin_requisicoes", **redirect_kwargs))

    try:
        horas_solicitadas = float(horas_raw)
        if horas_solicitadas < 0:
            raise ValueError()
    except Exception:
        flash("Informe um valor válido para horas solicitadas.", "error")
        return redirect(url_for("admin_requisicoes", **redirect_kwargs))

    if not data_evento:
        flash("Informe uma data válida para o evento.", "error")
        return redirect(url_for("admin_requisicoes", **redirect_kwargs))

    if not atividade_id:
        atividade_id = requisicao["atividade_id"]

    allowed_activity_ids, _matriz = get_allowed_activity_ids_for_turma_matrix(
        conn,
        requisicao["turma_curso_id"],
        requisicao["turma_matriz_id"],
    )
    current_atividade_id = requisicao["atividade_id"]
    if (
        _has_versioned_requisicao_snapshot(requisicao)
        and atividade_id != current_atividade_id
    ):
        flash(
            "".join(
                (
                    "Esta solicitação já possui versão normativa registrada. ",
                    "Para trocar a atividade, crie uma nova solicitação.",
                )
            ),
            "error",
        )
        return redirect(url_for("admin_requisicoes", **redirect_kwargs))
    if atividade_id != current_atividade_id and atividade_id not in allowed_activity_ids:
        flash("A atividade selecionada não pertence à matriz efetiva da turma do aluno.", "error")
        return redirect(url_for("admin_requisicoes", **redirect_kwargs))

    atividade = conn.execute("SELECT id FROM atividades WHERE id = ?", (atividade_id,)).fetchone()
    if not atividade:
        flash("Atividade não encontrada.", "error")
        return redirect(url_for("admin_requisicoes", **redirect_kwargs))

    arquivos = request.files.getlist("comprovantes_files") or []
    first_saved = _append_requisicao_arquivos(conn, req_id, requisicao["aluno_id"], arquivos)

    params = [atividade_id, nome_evento, horas_solicitadas, data_evento, observacao]
    sql = """
        UPDATE requisicoes
           SET atividade_id = ?,
               nome_evento = ?,
               horas_solicitadas = ?,
               data_evento = ?,
               observacao = ?
    """
    if requisicao["arquivo_comprovante"] is None and first_saved:
        sql += ", arquivo_comprovante = ?"
        params.append(first_saved)
    sql += " WHERE id = ?"
    params.append(req_id)

    conn.execute(sql, tuple(params))
    conn.commit()
    flash("Requisição atualizada com sucesso.", "success")
    return redirect(url_for("admin_requisicoes"))


@admin_required
def admin_excluir_requisicao(req_id):
    conn = get_db_connection()
    row = conn.execute("SELECT id FROM requisicoes WHERE id = ?", (req_id,)).fetchone()
    if not row:
        return ("Requisição não encontrada.", 404)
    try:
        conn.execute("DELETE FROM requisicao_arquivos WHERE requisicao_id = ?", (req_id,))
        conn.execute("DELETE FROM requisicoes WHERE id = ?", (req_id,))
        conn.commit()
    except Exception as e:
        logger.error(f"Erro ao excluir requisição {req_id}: {e}")
        return ("Erro ao excluir.", 500)
    # Remove arquivos físicos (best-effort)
    try:
        upload_root = current_app.config.get("UPLOAD_FOLDER")
        if upload_root:
            req_dir = os.path.join(upload_root, f"req_{req_id}")
            if os.path.isdir(req_dir):
                shutil.rmtree(req_dir, ignore_errors=True)
    except Exception:
        pass
    return ("", 204)


@admin_required
def admin_detalhes_requisicao(req_id):
    conn = get_db_connection()
    requisicao = conn.execute("""
        SELECT r.*, u.nome as AdminNome, a.nome as AlunoNome, a.matricula as AlunoMatricula, act.nome as AtividadeNome
        FROM requisicoes r
        LEFT JOIN usuarios u ON r.admin_id = u.id
        LEFT JOIN alunos a ON r.aluno_id = a.id
        JOIN atividades act ON r.atividade_id = act.id
        WHERE r.id = ?
    """, (req_id,)).fetchone()
    if not requisicao:
        flash("Requisição não encontrada.", "error")
        return redirect(url_for("admin_requisicoes"))
    return render_template("admin_detalhes_requisicao.html", requisicao=requisicao)


@admin_required
def admin_api_requisicao(req_id):
    """Retorna detalhes da requisição para hidratar o modal (inclui anexos)."""
    conn = get_db_connection()
    ensure_turmas_matriz_schema(conn)
    ensure_matriz_atividade_links_table(conn)
    r = conn.execute(
        """
        SELECT r.*, a.nome as aluno_nome, a.turma_id as turma_id,
             COALESCE(t.codigo, t.nome, 'Sem turma') as turma_label,
             t.curso_id as turma_curso_id, t.matriz_id as turma_matriz_id,
               act.nome as atividade_nome, act.grupo as grupo, act.tipo_atividade as tipo_atividade
          FROM requisicoes r
          LEFT JOIN alunos a ON r.aluno_id = a.id
          LEFT JOIN turmas t ON t.id = a.turma_id
          JOIN atividades act ON r.atividade_id = act.id
         WHERE r.id = ?
        """,
        (req_id,)
    ).fetchone()
    if not r:
        return jsonify({"error":"not-found"}), 404
    anexos = conn.execute(
        "SELECT id, label, filename, criado_em FROM requisicao_arquivos WHERE requisicao_id = ? ORDER BY id",
        (req_id,)
    ).fetchall()
    def row_to_dict(row):
        if row is None:
            return None
        try:
            return {k: row[k] for k in row.keys()}
        except Exception:
            # fallback: sqlite3.Row supports .keys(); if not, map by items
            return dict(row)
    data = row_to_dict(r)
    allowed_activity_ids, matriz = get_allowed_activity_ids_for_turma_matrix(
        conn,
        data.get("turma_curso_id"),
        data.get("turma_matriz_id"),
    )
    data["allowed_activity_ids"] = sorted(allowed_activity_ids)
    data["current_activity_allowed"] = data.get("atividade_id") in allowed_activity_ids
    data["matriz_scope"] = (
        {"id": matriz["id"], "label": _matriz_option_label(matriz)} if matriz else None
    )
    data["anexos"] = [row_to_dict(x) for x in anexos]
    # URL pública para cada anexo (se possível)
    base_upload = current_app.config.get("UPLOAD_FOLDER")
    items = []
    for x in data["anexos"]:
        fn = x.get("filename")
        if not fn:
            items.append(x); continue
        # normalizar caminho para rota /uploads
        safe = os.path.normpath(fn).replace("\\", "/").lstrip("/")
        x["url"] = url_for("uploaded_file", filename=safe)
        items.append(x)
    data["anexos"] = items
    return jsonify(data)


@admin_required
def admin_api_aluno_requisicao_scope(aluno_id):
    conn = get_db_connection()
    scope = _get_admin_requisicao_scope_for_aluno(conn, aluno_id)
    if not scope:
        return jsonify({"error": "not-found"}), 404
    aluno = scope["aluno"]
    return jsonify(
        {
            "aluno_id": aluno["id"],
            "aluno_nome": aluno["nome"],
            "aluno_matricula": aluno["matricula"],
            "turma_id": aluno["turma_id"],
            "turma_label": scope["turma_label"],
            "allowed_activity_ids": scope["allowed_activity_ids"],
            "matriz_scope": scope["matriz_scope"],
        }
    )


@admin_required
def admin_processar_requisicao(req_id):
    conn = get_db_connection()
    ensure_turmas_matriz_schema(conn)
    ensure_matriz_atividade_links_table(conn)
    snapshot_display_enabled = is_versioned_requisicao_snapshot_display_enabled()
    requisicao = conn.execute("""
        SELECT r.*, a.nome as atividade_nome, a.tipo_atividade AS atividade_tipo_legacy_atual,
               a.tem_limitacao, a.tipo_limitacao,
               a.limite_horas_total, a.limite_horas_semestral, al.nome as aluno_nome,
               al.id AS aluno_rel_id, al.turma_id AS aluno_turma_id,
               t.id AS turma_rel_id,
               t.curso_id AS turma_curso_id, t.matriz_id AS turma_matriz_id
        FROM requisicoes r
        JOIN atividades a ON r.atividade_id = a.id
        LEFT JOIN alunos al ON r.aluno_id = al.id
        LEFT JOIN turmas t ON t.id = al.turma_id
        WHERE r.id = ?
    """, (req_id,)).fetchone()

    if not requisicao:
        flash("Requisição não encontrada.", "error")
        return redirect(url_for("admin_requisicoes"))

    snapshot_processing = read_requisicao_snapshot_for_processing(requisicao)
    snapshot_diag = _build_admin_requisicao_snapshot_diagnostic(requisicao)

    if (
        request.method == "POST"
        and snapshot_processing.authority is SnapshotProcessingAuthority.INVALID_AUTHORITATIVE_SNAPSHOT
    ):
        flash(RequisicaoSnapshotError.user_message, "error")
        return redirect(url_for("admin_processar_requisicao", req_id=req_id))

    if request.method == "GET":
        return render_template(
            "admin_processar_requisicao.html",
            requisicao=requisicao,
            snapshot_diag=snapshot_diag,
            snapshot_display_enabled=snapshot_display_enabled,
        )

    if request.method == "POST":
        status = request.form["status"]
        horas_deferidas = request.form.get("horas_deferidas")
        observacao = request.form.get("observacao")
        admin_id = session["user_id"]
        data_processamento = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        allowed_statuses = {
            "Pendente",
            "Deferida",
            "Deferida Parcialmente",
            "Indeferida",
            "Devolvida",
            "Encerrada",
        }

        if status not in allowed_statuses:
            flash("Status de processamento inválido.", "error")
            return redirect(url_for("admin_processar_requisicao", req_id=req_id))

        if status == "Deferida Parcialmente" and (horas_deferidas is None or str(horas_deferidas).strip() == ""):
            flash("Horas deferidas são obrigatórias para status 'Deferida Parcialmente'.", "error")
            return redirect(url_for("admin_processar_requisicao", req_id=req_id))
        # normaliza horas_deferidas se necessário
        if status == "Deferida Parcialmente":
            try:
                horas_deferidas = float(horas_deferidas)
                if horas_deferidas < 0:
                    raise ValueError()
            except Exception:
                flash("Informe um número válido para horas deferidas.", "error")
                return redirect(url_for("admin_processar_requisicao", req_id=req_id))

        if status in ["Deferida", "Deferida Parcialmente"]:
            has_existing_turma = (
                requisicao["aluno_rel_id"] is not None
                and requisicao["turma_rel_id"] is not None
            )
            has_historical_null_matrix = (
                has_existing_turma and requisicao["turma_matriz_id"] is None
            )
            has_explicit_matrix = (
                has_existing_turma and requisicao["turma_matriz_id"] is not None
            )

            if not has_existing_turma:
                flash("Dados do aluno não encontrados.", "error")
                return redirect(url_for("admin_processar_requisicao", req_id=req_id))

            if (
                snapshot_processing.authority is not SnapshotProcessingAuthority.VALID_AUTHORITATIVE_SNAPSHOT
                and not has_historical_null_matrix
                and has_explicit_matrix
                and not is_activity_allowed_for_turma_matrix(
                    conn,
                    requisicao["atividade_id"],
                    requisicao["turma_curso_id"],
                    requisicao["turma_matriz_id"],
                )
            ):
                flash(
                    "A atividade desta requisição não pertence mais a matriz efetiva da turma do aluno.",
                    "error",
                )
                return redirect(url_for("admin_processar_requisicao", req_id=req_id))

        if status in ["Deferida", "Deferida Parcialmente"]:
            horas_a_deferir = float(horas_deferidas) if status == "Deferida Parcialmente" else float(requisicao["horas_solicitadas"])
            snapshot_rule = (
                snapshot_processing.rule
                if snapshot_processing.authority is SnapshotProcessingAuthority.VALID_AUTHORITATIVE_SNAPSHOT
                else None
            )
            total_limit = snapshot_rule.limite_total if snapshot_rule else None
            semester_limit = snapshot_rule.limite_semestre if snapshot_rule else None
            if snapshot_rule is None and requisicao["tem_limitacao"]:
                total_limit = requisicao["limite_horas_total"] if requisicao["tipo_limitacao"] == "total" else None
                semester_limit = requisicao["limite_horas_semestral"] if requisicao["tipo_limitacao"] == "semestral" else None

            if total_limit is not None:
                horas_ja_deferidas = conn.execute("""
                    SELECT COALESCE(SUM(
                        CASE\x20
                            WHEN status = 'Deferida' THEN horas_solicitadas
                            WHEN status = 'Deferida Parcialmente' THEN horas_deferidas
                            ELSE 0
                        END
                    ), 0) as total
                    FROM requisicoes\x20
                    WHERE aluno_id = ? AND atividade_id = ? AND status IN ('Deferida', 'Deferida Parcialmente')
                """, (requisicao["aluno_id"], requisicao["atividade_id"])).fetchone()[0]

                if horas_ja_deferidas + horas_a_deferir > total_limit:
                    flash(f"Erro: O aluno já possui {horas_ja_deferidas}h nesta atividade. Limite total: {total_limit}h. Máximo a deferir agora: {total_limit - horas_ja_deferidas}h.", "error")
                    return redirect(url_for("admin_processar_requisicao", req_id=req_id))

            if semester_limit is not None:
                ano_atual = datetime.datetime.now().year
                semestre_atual = 1 if datetime.datetime.now().month <= 6 else 2

                horas_ja_deferidas_semestre = conn.execute("""
                    SELECT COALESCE(SUM(
                        CASE\x20
                            WHEN status = 'Deferida' THEN horas_solicitadas
                            WHEN status = 'Deferida Parcialmente' THEN horas_deferidas
                            ELSE 0
                        END
                    ), 0) as total
                    FROM requisicoes\x20
                    WHERE aluno_id = ? AND atividade_id = ? AND status IN ('Deferida', 'Deferida Parcialmente')
                    AND strftime('%Y', data_evento) = ?\x20
                    AND (
                        (? = 1 AND strftime('%m', data_evento) BETWEEN '01' AND '06') OR
                        (? = 2 AND strftime('%m', data_evento) BETWEEN '07' AND '12')
                    )
                """, (requisicao["aluno_id"], requisicao["atividade_id"], str(ano_atual), semestre_atual, semestre_atual)).fetchone()[0]

                if horas_ja_deferidas_semestre + horas_a_deferir > semester_limit:
                    flash(f"Erro: Já possui {horas_ja_deferidas_semestre}h neste semestre. Limite semestral: {semester_limit}h. Máximo agora: {semester_limit - horas_ja_deferidas_semestre}h.", "error")
                    return redirect(url_for("admin_processar_requisicao", req_id=req_id))

        if status != "Deferida Parcialmente":
            horas_deferidas = None

        if status == "Pendente":
            data_processamento = None
            admin_id = None

        set_parts = [
            "status = ?",
            "horas_deferidas = ?",
            "observacao = ?",
            "data_processamento = ?",
            "admin_id = ?",
        ]
        params = [status, horas_deferidas, observacao, data_processamento, admin_id]
        current_status = str(requisicao["status"] or "").strip()
        if status == "Pendente":
            set_parts.extend([
                "aluno_update_notified_at = ?",
                "aluno_update_seen_at = ?",
            ])
            params.extend([None, None])
        elif current_status == "Pendente":
            set_parts.extend([
                "aluno_update_notified_at = ?",
                "aluno_update_seen_at = ?",
            ])
            params.extend([data_processamento, None])

        params.append(req_id)
        conn.execute(
            f"UPDATE requisicoes SET {', '.join(set_parts)} WHERE id = ?",
            params,
        )
        conn.commit()
        flash("Requisição processada com sucesso.", "success")
        return redirect(url_for("admin_requisicoes"))

    return render_template(
        "admin_processar_requisicao.html",
        requisicao=requisicao,
        snapshot_diag=snapshot_diag,
        snapshot_display_enabled=snapshot_display_enabled,
    )


bp_admin_requisicoes = Blueprint("admin_requisicoes_blueprint", __name__)

LEGACY_ROUTE_SPECS = configure_legacy_routes(
    bp_admin_requisicoes,
    (
        LegacyRouteSpec(
            "/admin/importar_requisicoes",
            "admin_importar_requisicoes",
            admin_importar_requisicoes,
            ("GET", "POST"),
        ),
        LegacyRouteSpec(
            "/admin/requisicoes",
            "admin_requisicoes",
            admin_requisicoes,
            ("GET",),
        ),
        LegacyRouteSpec(
            "/admin/requisicoes/nova",
            "admin_nova_requisicao",
            admin_nova_requisicao,
            ("GET", "POST"),
        ),
        LegacyRouteSpec(
            "/admin/requisicoes/<int:req_id>/editar",
            "admin_editar_requisicao",
            admin_editar_requisicao,
            ("POST",),
        ),
        LegacyRouteSpec(
            "/admin/requisicoes/<int:req_id>/excluir",
            "admin_excluir_requisicao",
            admin_excluir_requisicao,
            ("POST",),
        ),
        LegacyRouteSpec(
            "/admin/requisicao/<int:req_id>",
            "admin_detalhes_requisicao",
            admin_detalhes_requisicao,
            ("GET",),
        ),
        LegacyRouteSpec(
            "/admin/api/requisicao/<int:req_id>",
            "admin_api_requisicao",
            admin_api_requisicao,
            ("GET",),
        ),
        LegacyRouteSpec(
            "/admin/api/aluno/<int:aluno_id>/requisicao-scope",
            "admin_api_aluno_requisicao_scope",
            admin_api_aluno_requisicao_scope,
            ("GET",),
        ),
        LegacyRouteSpec(
            "/admin/processar_requisicao/<int:req_id>",
            "admin_processar_requisicao",
            admin_processar_requisicao,
            ("GET", "POST"),
        ),
    ),
)


__all__ = [
    "ALLOWED_EXCEL",
    "LEGACY_ROUTE_SPECS",
    "admin_api_aluno_requisicao_scope",
    "admin_api_requisicao",
    "admin_detalhes_requisicao",
    "admin_editar_requisicao",
    "admin_excluir_requisicao",
    "admin_importar_requisicoes",
    "admin_nova_requisicao",
    "admin_processar_requisicao",
    "admin_requisicoes",
    "bp_admin_requisicoes",
    "_append_requisicao_arquivos",
    "_get_admin_requisicao_scope_for_aluno",
    "_list_admin_requisicao_alunos",
    "_normalize_requisicao_data_evento",
]
